"""
Модуль для работы с базой данных
ВЕРСИЯ С ПОДДЕРЖКОЙ ГРУПП
"""

import sqlite3
import re
from datetime import datetime
from typing import List, Tuple, Dict
from config import CURRENCIES
import os
import logging
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

logger = logging.getLogger(__name__)

# БД лежит рядом с файлами бота
DB_PATH = os.path.join(os.path.dirname(__file__), "operations.db")


class Database:
    @staticmethod
    def _norm(s: str) -> str:
        s = (s or "").strip().lower()
        s = " ".join(s.split())
        return s
    def __init__(self, db_name: str = DB_PATH):
        """Инициализация базы данных"""
        self.db_name = db_name
        self.create_tables()
    # def get_all_chats(self):
        # self.cursor.execute("SELECT DISTINCT chat_id FROM operations")
        # return [row[0] for row in self.cursor.fetchall()]


    def get_report_income_by_date(self, chat_id: int, report_date: str):
        """
        Возвращает список строк для отчёта:
        [(client_name, currency, amount, full_message), ...]
        где client_name извлечен из description,
        а full_message — исходный description (пойдет в комментарий).
        """
        conn = self.get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                COALESCE(NULLIF(TRIM(description), ''), 'Без клиента') AS full_message,
                currency,
                amount
            FROM operations
            WHERE chat_id = ?
            AND amount > 0
            AND date(timestamp) = date(?)
            ORDER BY timestamp ASC
        """, (chat_id, report_date))

        rows = cur.fetchall()
        conn.close()

        agg = defaultdict(float)                 # (client_name, currency) -> sum
        msgs = defaultdict(list)                 # client_name -> list of full messages

        for r in rows:
            full_message = r["full_message"]
            cur_ = r["currency"]
            amt = float(r["amount"] or 0.0)

            client_name = self.extract_client_name(full_message)
            key = (client_name, cur_)
            agg[key] += amt

            if full_message:
                msgs[client_name].append(str(full_message))

        # собираем итоговый список
        out = []
        for (client_name, cur_), total_amt in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
            # в комментарий — все сообщения клиента за день (склеим)
            full_text = "\n\n---\n\n".join(msgs.get(client_name, []))
            out.append((client_name, cur_, float(total_amt), full_text))

        return out


    def get_connection(self):
        """Создание стабильного подключения к SQLite (safe for asyncio)"""
        conn = sqlite3.connect(
            self.db_name,
            timeout=30,
            check_same_thread=False,
            isolation_level=None,   # автокоммит, снижает lock'и
        )

        conn.row_factory = sqlite3.Row

        # ⚠️ PRAGMA — строго в таком порядке
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA busy_timeout=8000;")

        return conn


    def create_tables(self):
        """Создание таблиц в БД"""
        conn = self.get_connection()
        cursor = conn.cursor()

        # Таблица операций с chat_id
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id INTEGER NOT NULL,
                operation_type TEXT NOT NULL,
                currency TEXT NOT NULL,
                amount REAL NOT NULL,
                description TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Индекс для быстрого поиска по chat_id
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_operations_chat_id
            ON operations(chat_id)
        ''')

        # Таблица балансов с chat_id
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS balances (
                chat_id INTEGER NOT NULL,
                currency TEXT NOT NULL,
                balance REAL DEFAULT 0.0,
                last_updated DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (chat_id, currency)
            )
        ''')

        # Таблица с информацией о чатах
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS chats (
                chat_id INTEGER PRIMARY KEY,
                chat_name TEXT,
                chat_type TEXT,
                first_interaction DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_interaction DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # NEW: Таблица начальных остатков (Cash Evening Report)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cash_opening_balances (
                date TEXT NOT NULL,
                currency TEXT NOT NULL,
                amount REAL NOT NULL,
                group_id INTEGER DEFAULT 0,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (date, currency, group_id)
            )
        ''')

        # NEW: Таблица внутренних курсов (Cash Evening Report)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS internal_rates (
                group_id INTEGER DEFAULT 0,
                from_currency TEXT NOT NULL,
                to_currency TEXT NOT NULL,
                rate REAL NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (group_id, from_currency, to_currency)
            )
        ''')

        conn.commit()
        conn.close()

    def register_chat(self, chat_id: int, chat_name: str = None, chat_type: str = 'private'):
        """Регистрация чата/группы"""
        conn = self.get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT OR REPLACE INTO chats (chat_id, chat_name, chat_type, last_interaction)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        ''', (chat_id, chat_name, chat_type))

        # Инициализация балансов для всех валют для этого чата
        for currency in CURRENCIES:
            cursor.execute('''
                INSERT OR IGNORE INTO balances (chat_id, currency, balance)
                VALUES (?, ?, 0.0)
            ''', (chat_id, currency))

        conn.commit()
        conn.close()

    def add_operation(
        self,
        chat_id: int,
        operation_type: str,
        currency: str,
        amount: float,
        description: str = ""
    ) -> int:
        """
        Добавить операцию для конкретного чата

        Args:
            chat_id: ID чата/группы
            operation_type: Тип операции
            currency: Валюта
            amount: Сумма (положительная для прихода, отрицательная для расхода)
            description: Описание операции

        Returns:
            ID добавленной операции
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        # Убедимся что чат зарегистрирован
        cursor.execute('SELECT chat_id FROM chats WHERE chat_id = ?', (chat_id,))
        if not cursor.fetchone():
            cursor.execute('''
                INSERT INTO chats (chat_id) VALUES (?)
            ''', (chat_id,))
            # Инициализация балансов
            for curr in CURRENCIES:
                cursor.execute('''
                    INSERT OR IGNORE INTO balances (chat_id, currency, balance)
                    VALUES (?, ?, 0.0)
                ''', (chat_id, curr))

        # Добавляем операцию
        cursor.execute('''
            INSERT INTO operations (chat_id, operation_type, currency, amount, description)
            VALUES (?, ?, ?, ?, ?)
        ''', (chat_id, operation_type, currency, amount, description))

        operation_id = cursor.lastrowid

        # Обновляем баланс для этого чата
        cursor.execute('''
            INSERT INTO balances (chat_id, currency, balance, last_updated)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(chat_id, currency) DO UPDATE SET
                balance = balance + ?,
                last_updated = CURRENT_TIMESTAMP
        ''', (chat_id, currency, amount, amount))

        # Обновляем время последнего взаимодействия
        cursor.execute('''
            UPDATE chats SET last_interaction = CURRENT_TIMESTAMP WHERE chat_id = ?
        ''', (chat_id,))

        conn.commit()
        conn.close()

        return operation_id

    def get_balances(self, chat_id: int) -> Dict[str, float]:
        """
        Получить балансы для конкретного чата

        Args:
            chat_id: ID чата/группы

        Returns:
            Словарь {валюта: баланс}
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        # Убедимся что чат зарегистрирован в balances
        cursor.execute('SELECT chat_id FROM balances WHERE chat_id = ? LIMIT 1', (chat_id,))
        if not cursor.fetchone():
            # Инициализируем балансы
            for currency in CURRENCIES:
                cursor.execute('''
                    INSERT OR IGNORE INTO balances (chat_id, currency, balance)
                    VALUES (?, ?, 0.0)
                ''', (chat_id, currency))
            conn.commit()

        cursor.execute('''
            SELECT currency, balance
            FROM balances
            WHERE chat_id = ?
            ORDER BY currency
        ''', (chat_id,))
        rows = cursor.fetchall()

        conn.close()

        result = {row["currency"]: row["balance"] for row in rows}

        # Добавляем недостающие валюты из CURRENCIES
        for currency in CURRENCIES:
            if currency not in result:
                result[currency] = 0.0

        return result

    def get_balance(self, chat_id: int, currency: str) -> float:
        """
        Получить баланс по конкретной валюте для чата
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT balance FROM balances
            WHERE chat_id = ? AND currency = ?
        ''', (chat_id, currency))
        row = cursor.fetchone()

        conn.close()

        return row["balance"] if row else 0.0

    def get_group_balances_table(self) -> Dict[str, Dict[str, float]]:
        """
        Таблица остатков:
        {
          "Название группы": {"USD": 10, "RUB": 500, ...},
          ...
        }
        """
        conn = self.get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                COALESCE(c.chat_name, CAST(b.chat_id AS TEXT)) AS group_name,
                b.currency,
                b.balance
            FROM balances b
            LEFT JOIN chats c ON c.chat_id = b.chat_id
            ORDER BY group_name, b.currency
        """)
        rows = cur.fetchall()
        conn.close()

        table = defaultdict(dict)
        for r in rows:
            table[r["group_name"]][r["currency"]] = float(r["balance"] or 0.0)

        # гарантируем все валюты из CURRENCIES
        for group_name in table:
            for currency in CURRENCIES:
                table[group_name].setdefault(cur, 0.0)
        # после заполнения валют
        table = {g: curmap for g, curmap in table.items()
                if any(abs(v) > 1e-9 for v in curmap.values())}


        return dict(table)
    
    def get_total_balances_all_groups(self) -> Dict[str, float]:
        """Итого по валютам по всем чатам/группам"""
        conn = self.get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT currency, COALESCE(SUM(balance), 0) AS total
            FROM balances
            GROUP BY currency
            ORDER BY currency
        """)
        rows = cur.fetchall()
        conn.close()

        result = {r["currency"]: float(r["total"] or 0.0) for r in rows}
        for cur_ in CURRENCIES:
            result.setdefault(cur_, 0.0)
        return result

    def get_operations(
        self,
        chat_id: int,
        limit: int = 20,
        currency: str | None = None
    ) -> List[Tuple]:
        """
        Получить список операций для конкретного чата

        Returns:
            Список кортежей (id, operation_type, currency, amount, description, timestamp)
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        if currency:
            cursor.execute(
                '''
                SELECT id, operation_type, currency, amount, description,
                       strftime('%d.%m.%Y %H:%M', timestamp) as timestamp
                FROM operations
                WHERE chat_id = ? AND currency = ?
                ORDER BY id DESC
                LIMIT ?
                ''',
                (chat_id, currency, limit),
            )
        else:
            cursor.execute(
                '''
                SELECT id, operation_type, currency, amount, description,
                       strftime('%d.%m.%Y %H:%M', timestamp) as timestamp
                FROM operations
                WHERE chat_id = ?
                ORDER BY id DESC
                LIMIT ?
                ''',
                (chat_id, limit),
            )

        rows = cursor.fetchall()
        conn.close()

        return [
            (
                row["id"],
                row["operation_type"],
                row["currency"],
                row["amount"],
                row["description"],
                row["timestamp"],
            )
            for row in rows
        ]
    

    def get_operations_by_date(self, chat_id: int, date_from=None, date_to=None):
        conn = self.get_connection()
        cur = conn.cursor()

        logger.info(f"get_operations_by_date: chat_id={chat_id}, from={date_from}, to={date_to}")

        if date_from and date_to:
            # Конвертируем date в строку формата YYYY-MM-DD
            date_from_str = date_from.strftime("%Y-%m-%d") if hasattr(date_from, 'strftime') else str(date_from)
            date_to_str = date_to.strftime("%Y-%m-%d") if hasattr(date_to, 'strftime') else str(date_to)

            # logger.info(f"Запрос БД: FROM {date_from_str} TO {date_to_str}")

            cur.execute("""
                SELECT id, operation_type, currency, amount, description,
                    strftime('%d.%m.%Y %H:%M', timestamp) as timestamp
                FROM operations
                WHERE chat_id = ?
                AND date(timestamp) BETWEEN date(?) AND date(?)
                ORDER BY timestamp
            """, (chat_id, date_from_str, date_to_str))

        elif date_from:
            date_from_str = date_from.strftime("%Y-%m-%d") if hasattr(date_from, 'strftime') else str(date_from)

            logger.info(f"Запрос БД: DATE {date_from_str}")

            cur.execute("""
                SELECT id, operation_type, currency, amount, description,
                    strftime('%d.%m.%Y %H:%M', timestamp) as timestamp
                FROM operations
                WHERE chat_id = ?
                AND date(timestamp) = date(?)
                ORDER BY timestamp
            """, (chat_id, date_from_str))

        else:
            logger.info("Запрос БД: ВСЕ операции")

            cur.execute("""
                SELECT id, operation_type, currency, amount, description,
                    strftime('%d.%m.%Y %H:%M', timestamp) as timestamp
                FROM operations
                WHERE chat_id = ?
                ORDER BY timestamp
            """, (chat_id,))

        rows = cur.fetchall()
        logger.info(f"Найдено операций: {len(rows)}")

        conn.close()

        return [
            (
                row["id"],
                row["operation_type"],
                row["currency"],
                row["amount"],
                row["description"],
                row["timestamp"],
            )
            for row in rows
        ]



    def get_statistics(self, chat_id: int) -> Dict[str, Dict[str, float]]:
        """
        Получить статистику для конкретного чата

        Returns:
            {валюта: {income, expense, balance}}
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        stats: Dict[str, Dict[str, float]] = {}

        for currency in CURRENCIES:
            # Сумма поступлений
            cursor.execute(
                '''
                SELECT COALESCE(SUM(amount), 0) as income
                FROM operations
                WHERE chat_id = ? AND currency = ? AND amount > 0
                ''',
                (chat_id, currency),
            )
            income = cursor.fetchone()["income"]

            # Сумма расходов
            cursor.execute(
                '''
                SELECT COALESCE(SUM(amount), 0) as expense
                FROM operations
                WHERE chat_id = ? AND currency = ? AND amount < 0
                ''',
                (chat_id, currency),
            )
            expense = cursor.fetchone()["expense"]

            # Текущий баланс
            balance = self.get_balance(chat_id, currency)

            if income != 0 or expense != 0 or balance != 0:
                stats[currency] = {
                    "income": income,
                    "expense": expense,
                    "balance": balance,
                }

        conn.close()
        return stats

    def get_total_operations_count(self, chat_id: int) -> int:
        """Получить общее количество операций для чата"""
        conn = self.get_connection()
        cursor = conn.cursor()

        cursor.execute(
            '''
            SELECT COUNT(*) as count
            FROM operations
            WHERE chat_id = ?
            ''',
            (chat_id,),
        )
        count = cursor.fetchone()["count"]

        conn.close()
        return count

    def get_all_chats(self) -> List[Tuple]:
        """Получить список всех чатов"""
        conn = self.get_connection()
        cursor = conn.cursor()

        cursor.execute(
            '''
            SELECT chat_id, chat_name, chat_type,
                   strftime('%d.%m.%Y %H:%M', first_interaction) as first_interaction,
                   strftime('%d.%m.%Y %H:%M', last_interaction) as last_interaction
            FROM chats
            ORDER BY last_interaction DESC
            '''
        )
        rows = cursor.fetchall()

        conn.close()
        return [
            (
                row["chat_id"],
                row["chat_name"],
                row["chat_type"],
                row["first_interaction"],
                row["last_interaction"],
            )
            for row in rows
        ]
    def get_chat(self, chat_id: int):
        """Получить один чат по chat_id (формат как get_all_chats)"""
        conn = self.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT chat_id, chat_name, chat_type,
                strftime('%d.%m.%Y %H:%M', first_interaction) as first_interaction,
                strftime('%d.%m.%Y %H:%M', last_interaction) as last_interaction
            FROM chats
            WHERE chat_id = ?
        """, (chat_id,))

        row = cursor.fetchone()
        conn.close()

        if not row:
            return None

        return (
            row["chat_id"],
            row["chat_name"],
            row["chat_type"],
            row["first_interaction"],
            row["last_interaction"],
        )

    
    @staticmethod
    def extract_client_name(text: str) -> str:
        """
        Пример текста:
        "...-Плательщик ООО \"АВТОЦЕНТРГАЗ-РУСАВТО\"- ЕВРО АВТО"
        Вернет: "ЕВРО АВТО"

        Логика:
        1) берем хвост после последнего дефиса (- или —) ближе к концу
        2) чистим пробелы/переводы строк
        """
        if not text:
            return "Без клиента"

        t = " ".join(str(text).split())  # нормализация пробелов/переносов

        # хвост после последнего " - " или "—" в конце
        m = re.search(r"(?:\s*[-—]\s*)([^-—]{2,})\s*$", t)
        if m:
            name = m.group(1).strip()
            return name or "Без клиента"

        return "Без клиента"

    def delete_operation(self, chat_id: int, operation_id: int) -> bool:
        """
        Удалить операцию (с откатом баланса)

        Returns:
            True если удалена, False если не найдена
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        # Получаем данные операции
        cursor.execute(
            '''
            SELECT currency, amount FROM operations
            WHERE id = ? AND chat_id = ?
            ''',
            (operation_id, chat_id),
        )
        row = cursor.fetchone()

        if not row:
            conn.close()
            return False

        currency = row["currency"]
        amount = row["amount"]

        # Удаляем операцию
        cursor.execute(
            "DELETE FROM operations WHERE id = ? AND chat_id = ?",
            (operation_id, chat_id),
        )

        # Откатываем баланс
        cursor.execute(
            '''
            UPDATE balances
            SET balance = balance - ?,
                last_updated = CURRENT_TIMESTAMP
            WHERE chat_id = ? AND currency = ?
            ''',
            (amount, chat_id, currency),
        )

        conn.commit()
        conn.close()

        return True

    @staticmethod
    def parse_timestamp(ts) -> datetime:
        """
        Универсальный парсер времени под SQLite CURRENT_TIMESTAMP.
        Ожидаемый основной формат: 'YYYY-MM-DD HH:MM:SS'
        + fallback для ISO и strftime '%d.%m.%Y %H:%M'
        """
        # уже datetime → возвращаем как есть
        if isinstance(ts, datetime):
            return ts

        # None/пустое → текущее
        if not ts:
            return datetime.now()

        # основной формат SQLite
        try:
            return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        except Exception:
            pass

        # ISO 8601 ('2025-01-03T15:14:22')
        try:
            return datetime.fromisoformat(ts.replace("Z", ""))
        except Exception:
            pass

        # формат strftime('%d.%m.%Y %H:%M')
        try:
            return datetime.strptime(ts, "%d.%m.%Y %H:%M")
        except Exception:
            pass

        # формат strftime('%d.%m.%Y %H:%M:%S')
        try:
            return datetime.strptime(ts, "%d.%m.%Y %H:%M:%S")
        except Exception:
            pass

        # на крайний случай — вернём now, чтобы приложение не падало
        return datetime.now()

    def get_chat_id_by_name(self, name: str) -> int | None:
        """Ищем chat_id по имени группы максимально устойчиво."""
        wanted = self._norm(name)
        if not wanted:
            return None

        chats = self.get_all_chats()  # [(chat_id, chat_name, chat_type, fi, li), ...]
        best_id = None
        best_score = 0

        for (cid, chat_name, _chat_type, _fi, _li) in chats:
            n = self._norm(chat_name)

            # 1) точное совпадение
            if n == wanted:
                return cid

            # 2) вхождение
            score = 0
            if wanted in n:
                score = 100 + len(wanted)
            else:
                # 3) пересечение по словам
                w_words = set(wanted.split())
                n_words = set(n.split())
                common = len(w_words & n_words)
                if common:
                    score = 10 * common

            if score > best_score:
                best_score = score
                best_id = cid

        return best_id if best_score >= 20 else None

    @staticmethod
    def safe_sheet_name(name: str, fallback: str) -> str:
        if not name:
            name = fallback
        for ch in r'\/:*?[]':
            name = name.replace(ch, "_")
        return name[:31].strip()


    def clear_all(self):
        conn = self.get_connection()
        cur = conn.cursor()

        cur.execute("DELETE FROM operations;")
        cur.execute("DELETE FROM chats;")

        conn.commit()
        conn.close()

    def recalculate_balances(self, chat_id: int | None = None):
        """
        Пересчитать балансы

        Args:
            chat_id: ID конкретного чата или None для всех чатов
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        if chat_id is not None:
            chats = [(chat_id,)]
        else:
            cursor.execute("SELECT DISTINCT chat_id FROM operations")
            chats = cursor.fetchall()

        for (cid,) in chats:
            # Обнуляем балансы
            cursor.execute(
                '''
                UPDATE balances SET balance = 0.0
                WHERE chat_id = ?
                ''',
                (cid,),
            )

            # Пересчитываем для каждой валюты
            for currency in CURRENCIES:
                cursor.execute(
                    '''
                    SELECT COALESCE(SUM(amount), 0) as total
                    FROM operations
                    WHERE chat_id = ? AND currency = ?
                    ''',
                    (cid, currency),
                )
                total = cursor.fetchone()["total"]

                cursor.execute(
                    '''
                    INSERT INTO balances (chat_id, currency, balance, last_updated)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(chat_id, currency) DO UPDATE SET
                        balance = ?,
                        last_updated = CURRENT_TIMESTAMP
                    ''',
                    (cid, currency, total, total),
                )

        conn.commit()
        conn.close()


    def export_group_balances_to_excel(self, filepath: str):
        """
        Экспорт таблицы остатков групп в Excel
        """
        table = self.get_group_balances_table()
        totals = self.get_total_balances_all_groups()
        currencies = list(CURRENCIES)

        wb = Workbook()
        ws = wb.active
        ws.title = "Остатки групп"

        # ---------- Заголовок ----------
        headers = ["Группа"] + currencies
        ws.append(headers)

        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = header_font

        # ---------- Данные по группам ----------
        for group_name in sorted(table.keys()):
            safe_name = str(group_name).replace("\n", " ").replace("|", "/")
            row = [safe_name] + [table[group_name].get(cur, 0.0) for cur in currencies]
            ws.append(row)

        # ---------- ИТОГО ----------
        total_row_idx = ws.max_row + 1
        ws.append(["ИТОГО"] + [totals.get(cur, 0.0) for cur in currencies])

        for col in range(1, len(headers) + 1):
            ws.cell(row=total_row_idx, column=col).font = Font(bold=True)

        # ---------- Форматирование ----------
        for col in range(2, len(headers) + 1):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = "#,##0.00"

        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        wb.save(filepath)

    def set_cash_opening_balance(self, date_str: str, currency: str, amount: float, group_id: int = 0):
        """
        Установить начальный остаток для кассы
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT OR REPLACE INTO cash_opening_balances (date, currency, amount, group_id, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (date_str, currency, amount, group_id))
        
        conn.commit()
        conn.close()

    def get_cash_opening_balances(self, date_str: str, group_id: int = 0) -> Dict[str, float]:
        """
        Получить все начальные остатки на дату
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT currency, amount FROM cash_opening_balances
            WHERE date = ? AND group_id = ?
        ''', (date_str, group_id))
        
        rows = cursor.fetchall()
        conn.close()
        
        return {row["currency"]: row["amount"] for row in rows}

    def set_internal_rate(self, from_curr: str, to_curr: str, rate: float, group_id: int = 0):
        """
        Установить внутренний курс
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT OR REPLACE INTO internal_rates (group_id, from_currency, to_currency, rate, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (group_id, from_curr, to_curr, rate))
        
        conn.commit()
        conn.close()

    def get_internal_rate(self, from_curr: str, to_curr: str, group_id: int = 0) -> float | None:
        """
        Получить внутренний курс. 
        Если прямого курса нет, можно попробовать обратный (1/rate), 
        но пока реализуем только прямой поиск.
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT rate FROM internal_rates
            WHERE group_id = ? AND from_currency = ? AND to_currency = ?
        ''', (group_id, from_curr, to_curr))
        
        row = cursor.fetchone()
        conn.close()
        
        return row["rate"] if row else None

