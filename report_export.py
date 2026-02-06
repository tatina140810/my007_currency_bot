from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

def _autosize(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 2, 45)

def export_report_income_matrix(rows, output_path: str, report_date: str):
    """
    rows: [(client_name, currency, amount, full_message), ...]
    full_message будет сохранён в комментарий к ячейке с клиентом.
    """
    agg = defaultdict(lambda: defaultdict(float))   # client -> currency -> sum
    totals = defaultdict(float)                     # currency -> sum
    comments = defaultdict(list)                    # client -> [full_message...]

    for client_name, cur, amt, full_msg in rows:
        agg[client_name][cur] += float(amt)
        totals[cur] += float(amt)
        if full_msg:
            comments[client_name].append(str(full_msg))

    currencies = sorted(totals.keys())
    clients = sorted(agg.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = f"Report_{report_date}"

    ws["A1"] = "Клиент"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    for j, cur in enumerate(currencies, start=2):
        c = ws.cell(row=1, column=j, value=cur)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    r = 2
    for client in clients:
        cell_a = ws.cell(row=r, column=1, value=client)

        # ✅ Комментарий к ячейке клиента
        if comments.get(client):
            note_text = "\n\n---\n\n".join(comments[client])

            # ограничение Excel на комментарии — подстрахуемся
            if len(note_text) > 30000:
                note_text = note_text[:30000] + "\n\n... (обрезано)"

            cell_a.comment = Comment(note_text, "report-bot")

        for j, cur in enumerate(currencies, start=2):
            v = agg[client].get(cur, 0.0)
            cell = ws.cell(row=r, column=j, value=v if v != 0 else "")
            cell.number_format = "#,##0.00"
        r += 1

    ws.cell(row=r, column=1, value="ИТОГО").font = Font(bold=True)
    for j, cur in enumerate(currencies, start=2):
        cell = ws.cell(row=r, column=j, value=totals.get(cur, 0.0))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"

    _autosize(ws)
    wb.save(output_path)