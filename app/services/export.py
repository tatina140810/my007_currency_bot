"""
Модуль для экспорта операций в Excel
"""

import os
import logging
from collections import defaultdict
from datetime import datetime, date
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from app.db.database import Database
from app.core.config import CURRENCIES, OPERATION_TYPES

logger = logging.getLogger(__name__)

# ---------- разбор timestamp (дата, без часового пояса в Excel) ----------

def parse_timestamp(ts: str) -> datetime:
    """Пробуем разобрать разные форматы времени из БД."""
    if isinstance(ts, datetime):
        return ts

    if not ts:
        return datetime.now()

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(ts, fmt)
        except ValueError:
            continue

    return datetime.now()


def _date_key_from_timestamp(ts) -> str:
    """Ключ дня в формате dd.mm.yyyy."""
    dt = parse_timestamp(ts)
    return dt.strftime("%d.%m.%Y")


# ---------- стили ----------

def _create_styles() -> Dict[str, object]:
    """Создаём и возвращаем словарь со всеми используемыми стилями."""
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    income_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    expense_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    conversion_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    summary_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")

    return {
        "header_font": header_font,
        "header_fill": header_fill,
        "header_alignment": header_alignment,
        "income_fill": income_fill,
        "expense_fill": expense_fill,
        "conversion_fill": conversion_fill,
        "summary_header_fill": summary_header_fill,
        "section_fill": section_fill,
    }


# ---------- разбор суммы SWIFT из описания ----------

def _parse_swift_from_description(description: str) -> float:
    """
    Пробуем вытащить сумму SWIFT из текста описания, если пользователь её туда писал.
    """
    if not description:
        return 0.0

    import re
    pattern = r"(?i)(swift|свифт)[^\d\-]*([0-9]+(?:[.,][0-9]+)?)"
    m = re.search(pattern, description)
    if not m:
        return 0.0

    num_str = m.group(2).replace(",", ".")
    try:
        return float(num_str)
    except ValueError:
        return 0.0


# ---------- строки для таблицы Конвертация ----------

def _build_conversion_rows(conv_ops: List[Tuple]) -> List[List[object]]:
    rows: List[List[object]] = []
    i = 0
    n = len(conv_ops)

    while i < n:
        if i + 1 < n:
            op1 = conv_ops[i]
            op2 = conv_ops[i + 1]

            # buy = плюс, pay = минус
            if op1[3] > 0 and op2[3] < 0:
                buy_op, pay_op = op1, op2
            elif op2[3] > 0 and op1[3] < 0:
                buy_op, pay_op = op2, op1
            else:
                # если пара "сломана" — попробуем как есть
                buy_op, pay_op = op1, op2

            dt = parse_timestamp(buy_op[5] or pay_op[5])
            date_str = dt.strftime("%d.%m.%Y")

            buy_curr = buy_op[2]
            buy_amt = abs(buy_op[3])

            pay_curr = pay_op[2]
            pay_amt = abs(pay_op[3])

            rate = round(pay_amt / buy_amt, 6) if buy_amt else None

            rows.append([date_str, buy_amt, buy_curr, rate, pay_amt, pay_curr])

            i += 2
        else:
            # Одинокая конвертация (на всякий случай)
            op = conv_ops[i]
            dt = parse_timestamp(op[5])
            date_str = dt.strftime("%d.%m.%Y")

            if op[3] > 0:
                rows.append([date_str, abs(op[3]), op[2], None, "", ""])
            else:
                rows.append([date_str, "", "", None, abs(op[3]), op[2]])

            i += 1

    return rows


def _build_payment_maps(operations_sorted: List[Tuple]) -> tuple:
    swift_map: Dict[int, float] = {}
    commission_map: Dict[int, float] = {}

    n = len(operations_sorted)
    for i, op in enumerate(operations_sorted):
        op_id, op_type, currency, amount, description, timestamp = op
        if op_type != "Оплата ПП":
            continue

        swift_amount = 0.0
        commission_amount = 0.0
        payment_currency = currency

        for j in range(1, 3):
            if i + j >= n:
                break

            n_id, n_type, n_curr, n_amount, n_descr, n_ts = operations_sorted[i + j]

            if n_type == "Комиссия 1%" and n_curr == payment_currency:
                commission_amount = abs(n_amount)
            elif n_type == "SWIFT" and n_curr == "USD":
                swift_amount = abs(n_amount)

        swift_map[op_id] = swift_amount
        commission_map[op_id] = commission_amount

    return swift_map, commission_map


# ---------- ВЫРАВНИВАНИЕ ПО ДАТАМ ----------

def _group_rows_by_date(rows: List[List[object]], date_col_index: int = 1) -> dict:
    m = defaultdict(list)
    for r in rows:
        if len(r) <= date_col_index:
            continue
        d = r[date_col_index] or ""
        if d:
            m[d].append(r)
    return dict(m)


def _build_date_grid(rows_by_type: Dict[str, dict]) -> tuple[list[str], dict[str, int]]:
    all_dates = set()
    for dmap in rows_by_type.values():
        all_dates.update(dmap.keys())

    def _parse(d: str):
        try:
            return datetime.strptime(d, "%d.%m.%Y")
        except ValueError:
            return datetime.min

    all_dates_sorted = sorted(all_dates, key=_parse)

    span_by_date: dict[str, int] = {}
    for d in all_dates_sorted:
        span = 1
        for t in rows_by_type:
            span = max(span, len(rows_by_type[t].get(d, [])))
        span_by_date[d] = max(1, span)

    return all_dates_sorted, span_by_date


# ---------- Комментарии (описания) для ВСЕХ таблиц ----------

def _build_comment_map_for_type(ops_of_type: List[Tuple]) -> Dict[Tuple[str, int], str]:
    by_date = defaultdict(list)
    for op in ops_of_type:
        date_str = _date_key_from_timestamp(op[5])
        by_date[date_str].append(op)

    for d in by_date:
        by_date[d].sort(key=lambda o: parse_timestamp(o[5]))

    comment_map: Dict[Tuple[str, int], str] = {}
    for d, ops in by_date.items():
        for i, op in enumerate(ops):
            descr = (op[4] or "").strip()
            if descr and descr != "-":
                comment_map[(d, i)] = descr
    return comment_map


def _set_comment(cell, text: str):
    if not text:
        return
    txt = text.strip()
    if not txt:
        return
    cmt = Comment(txt, "system")
    cmt.width = 320
    cmt.height = 120
    cell.comment = cmt


# ---------- запись одного листа (один чат) ----------

def _write_operations_tables_for_chat(ws, operations: list, styles: dict, chat_id: int, db: Database):
    header_font = styles["header_font"]
    header_fill = styles["header_fill"]
    header_alignment = styles["header_alignment"]
    income_fill = styles["income_fill"]
    expense_fill = styles["expense_fill"]
    conversion_fill = styles["conversion_fill"]
    summary_header_fill = styles["summary_header_fill"]
    section_fill = styles["section_fill"]

    if not operations:
        ws["A1"] = "Нет операций для экспорта"
        ws["A1"].font = Font(bold=True)
        return

    # Сортируем по времени: старые сверху
    operations_sorted = sorted(operations, key=lambda op: parse_timestamp(op[5]))

    # Карты для SWIFT и Комиссии (привязка к Оплата ПП)
    swift_map, commission_map = _build_payment_maps(operations_sorted)

    # Группируем по типу операции (кроме SWIFT и Комиссия 1% — они привязаны к ПП)
    ops_by_type: Dict[str, List[Tuple]] = {}
    for op in operations_sorted:
        op_id, op_type, currency, amount, description, timestamp = op
        if op_type in ("SWIFT", "Комиссия 1%"):
            continue
        ops_by_type.setdefault(op_type, []).append(op)

    # Порядок типов: сначала OPERATION_TYPES, потом остальные
    ordered_types: List[str] = list(OPERATION_TYPES)
    extra_types = [t for t in ops_by_type.keys() if t not in ordered_types]
    ordered_types.extend(extra_types)

    # Комментарии по типам (описания)
    comment_map_by_type: Dict[str, Dict[Tuple[str, int], str]] = {}
    for t, ops_list in ops_by_type.items():
        comment_map_by_type[t] = _build_comment_map_for_type(ops_list)

    prepared_tables = []

    for op_type in ordered_types:
        if op_type not in ops_by_type:
            continue
        ops_list = ops_by_type[op_type]

        # --------- таблицы ---------

        if op_type == "Конвертация":
            headers = [
                "Дата",
                "Сумма откупа",
                "Валюта откупа",
                "Курс клиента",
                "Сумма оплаты",
                "Валюта оплаты",
            ]
            data_rows = _build_conversion_rows(ops_by_type[op_type])


        elif op_type == "Оплата ПП":
            headers = ["№", "Дата", "Валюта", "Сумма", "SWIFT USD", "Комиссия 1%"]
            data_rows = []
            idx = 1
            for op_id, t_op_type, currency, amount, description, timestamp in ops_by_type[op_type]:
                date_str = _date_key_from_timestamp(timestamp)

                swift_usd = swift_map.get(op_id, 0.0)
                commission = commission_map.get(op_id, 0.0)
                if not swift_usd:
                    swift_usd = _parse_swift_from_description(description)

                data_rows.append([idx, date_str, currency, amount, swift_usd, commission])
                idx += 1

        elif op_type == "Запрос банку":
            headers = ["№", "Дата", "Валюта", "Сумма (USD)"]
            data_rows = []
            idx = 1
            for op_id, t_op_type, currency, amount, description, timestamp in ops_by_type[op_type]:
                date_str = _date_key_from_timestamp(timestamp)
                data_rows.append([idx, date_str, currency, amount])
                idx += 1

        else:
            headers = ["№", "Дата", "Валюта", "Сумма"]
            data_rows = []
            idx = 1
            for op_id, t_op_type, currency, amount, description, timestamp in ops_by_type[op_type]:
                date_str = _date_key_from_timestamp(timestamp)
                data_rows.append([idx, date_str, currency, amount])
                idx += 1
        date_col_index = 0 if op_type == "Конвертация" else 1

        prepared_tables.append({
            "op_type": op_type,
            "headers": headers,
            "rows_by_date": _group_rows_by_date(data_rows, date_col_index=date_col_index),
            "cols_count": len(headers),
        })

    if not prepared_tables:
        ws["A1"] = "Нет операций для экспорта"
        ws["A1"].font = Font(bold=True)
        return

    rows_by_type_for_grid = {t["op_type"]: t["rows_by_date"] for t in prepared_tables}
    all_dates_sorted, span_by_date = _build_date_grid(rows_by_type_for_grid)

    total_rows = sum(span_by_date[d] for d in all_dates_sorted)
    if total_rows <= 0:
        ws["A1"] = "Нет операций для экспорта"
        ws["A1"].font = Font(bold=True)
        return

    start_col = 1

    for table in prepared_tables:
        op_type = table["op_type"]
        headers = table["headers"]
        rows_by_date = table["rows_by_date"]
        cols_count = table["cols_count"]

        # ----- строка 1: название секции -----
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=start_col + cols_count - 1,
        )
        section_cell = ws.cell(row=1, column=start_col, value=op_type)
        section_cell.font = Font(bold=True, size=13, color="FFFFFF")
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ----- строка 2: заголовки -----
        for offset, header in enumerate(headers):
            col_idx = start_col + offset
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

            col_letter = get_column_letter(col_idx)
            if header == "№":
                ws.column_dimensions[col_letter].width = 5
            elif header == "Дата":
                ws.column_dimensions[col_letter].width = 11
            elif header in ("Валюта", "Валюта (нал)", "Валюта платежа"):
                ws.column_dimensions[col_letter].width = 12
            elif header in ("Сумма", "Сумма (нал)", "Сумма платежа", "Сумма (USD)"):
                ws.column_dimensions[col_letter].width = 15
            elif header in ("SWIFT USD", "Комиссия 1%"):
                ws.column_dimensions[col_letter].width = 13
            elif header in ("Курс", "Курс клиента"):
                ws.column_dimensions[col_letter].width = 12
            elif header in ("Сумма откупа", "Сумма оплаты"):
                ws.column_dimensions[col_letter].width = 15
            elif header in ("Валюта откупа", "Валюта оплаты"):
                ws.column_dimensions[col_letter].width = 12
            elif header in ("Запрос банку"):
               ws.column_dimensions[col_letter].width = 15
            else:
                ws.column_dimensions[col_letter].width = 15

        # ----- данные с 3-й строки, выровненные по датам -----
        excel_row = 3
        for d in all_dates_sorted:
            span = span_by_date[d]
            day_rows = rows_by_date.get(d, [])

            for k in range(span):
                row_values = day_rows[k] if k < len(day_rows) else [""] * cols_count

                ws.row_dimensions[excel_row].height = 18

                descr = comment_map_by_type.get(op_type, {}).get((d, k), "")

                for offset in range(cols_count):
                    col_idx = start_col + offset
                    value = row_values[offset] if offset < len(row_values) else ""
                    cell = ws.cell(row=excel_row, column=col_idx, value=value)

                    header_text = headers[offset]
                    cell.alignment = Alignment(vertical="top", wrap_text=False)

                    # Форматы чисел
                    if header_text in (
                        "Сумма", "Сумма (нал)", "Сумма платежа", "Сумма (USD)",
                        "SWIFT USD", "Комиссия 1%",
                        "Сумма откупа", "Сумма оплаты",
                    ):
                        if isinstance(value, (int, float)):
                            cell.number_format = "#,##0.00"
                    elif header_text in ("Курс", "Курс клиента"):
                        if isinstance(value, (int, float)):
                            cell.number_format = "#,##0.000000"

                    if descr:
                        if op_type == "Конвертация" and header_text == "Сумма оплаты":
                            _set_comment(cell, descr)
                        elif op_type == "Запрос банку" and header_text == "Сумма (USD)":
                            _set_comment(cell, descr)
                        elif op_type != "Конвертация" and header_text == "Сумма":
                            _set_comment(cell, descr)

                if op_type == "Конвертация":
                    for off in range(cols_count):
                        ws.cell(row=excel_row, column=start_col + off).fill = conversion_fill
                else:
                    amount_idx = None
                    for idx_h, h in enumerate(headers):
                        if h.startswith("Сумма"):
                            amount_idx = idx_h
                            break

                    if amount_idx is not None and amount_idx < len(row_values):
                        amount_val = row_values[amount_idx]
                        if isinstance(amount_val, (int, float)):
                            row_fill = income_fill if amount_val > 0 else expense_fill
                            for off in range(cols_count):
                                ws.cell(row=excel_row, column=start_col + off).fill = row_fill

                excel_row += 1

        start_col = start_col + cols_count + 2

    ws.freeze_panes = "A3"

    # ---------- итоги по валютам ----------
    summary_start_row = 3 + total_rows + 2
    ws.cell(row=summary_start_row, column=1, value="ИТОГО ПО ВАЛЮТАМ:").font = Font(bold=True, size=12)
    summary_start_row += 1

    summary_headers = ["Валюта", "Поступления", "Расходы", "Баланс"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=summary_start_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = summary_header_fill
        cell.alignment = header_alignment
    summary_start_row += 1

    stats = db.get_statistics(chat_id)

    for currency in CURRENCIES:
        if currency in stats:
            currency_stats = stats[currency]

            ws.cell(row=summary_start_row, column=1, value=currency)

            income_cell = ws.cell(row=summary_start_row, column=2, value=currency_stats["income"])
            income_cell.number_format = "#,##0.00"
            income_cell.fill = income_fill

            expense_cell = ws.cell(row=summary_start_row, column=3, value=currency_stats["expense"])
            expense_cell.number_format = "#,##0.00"
            expense_cell.fill = expense_fill

            balance_cell = ws.cell(row=summary_start_row, column=4, value=currency_stats["balance"])
            balance_cell.number_format = "#,##0.00"
            if currency_stats["balance"] > 0:
                balance_cell.font = Font(bold=True, color="006100")
            elif currency_stats["balance"] < 0:
                balance_cell.font = Font(bold=True, color="9C0006")

            summary_start_row += 1


def export_to_excel(
    db: Database,
    output_path: str = os.path.join("outputs", "operations.xlsx"),
    date_from: date | None = None,
    date_to: date | None = None,
) -> str:
    """Экспорт всех операций всех чатов в одну книгу Excel"""

    logger.info(f"export_to_excel START: path={output_path}, from={date_from}, to={date_to}")

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Директория создана/проверена: {output_dir}")
    else:
        output_path = os.path.join("outputs", output_path)
        os.makedirs("outputs", exist_ok=True)
        logger.info(f"Используем путь: {output_path}")

    wb = Workbook()
    wb.remove(wb.active)

    styles = _create_styles()
    chats = db.get_all_chats()

    logger.info(f"Всего чатов в БД: {len(chats)}")

    if not chats:
        ws = wb.create_sheet("Нет данных")
        ws["A1"] = "Нет операций"
        ws["A1"].font = Font(bold=True)
        wb.save(output_path)
        logger.info("Сохранён пустой файл")
        return output_path

    sheets_created = 0

    for chat_id, chat_name, chat_type, *_ in chats:
        logger.info(f"Обработка чата: {chat_id} ({chat_name})")

        if date_from or date_to:
            operations = db.get_operations_by_date(chat_id, date_from, date_to)
        else:
            operations = db.get_operations(chat_id, limit=100000)

        logger.info(f"  Операций найдено: {len(operations)}")

        if not operations:
            logger.info(f"  Пропускаем чат {chat_id} (нет операций)")
            continue

        base_name = chat_name.strip() if chat_name else (
            "Группа" if chat_type in ("group", "supergroup") else "Чат"
        )

        sheet_name = f"{base_name}"[:31]
        for c in r'\/:*?[]':
            sheet_name = sheet_name.replace(c, "_")

        logger.info(f"  Создаём лист: {sheet_name}")

        ws = wb.create_sheet(sheet_name)
        _write_operations_tables_for_chat(ws, operations, styles, chat_id, db)

        sheets_created += 1
        logger.info(f"  Лист создан успешно")

    if sheets_created == 0:
        logger.info("Ни одного листа не создано, добавляем пустой")
        ws = wb.create_sheet("Нет данных")
        if date_from:
            ws["A1"] = f"Нет операций за {date_from.strftime('%d.%m.%Y')}"
        else:
            ws["A1"] = "Нет операций"
        ws["A1"].font = Font(bold=True)

    wb.save(output_path)
    logger.info(f"✅ Файл сохранён: {output_path}, листов: {sheets_created}")
    return output_path


def export_to_excel_single_chat(
    db: Database,
    chat_id: int,
    chat_name: str = None,
    output_path: str = os.path.join("outputs", "operations_single_chat.xlsx"),
) -> str:
    """Экспорт операций одного чата в отдельный файл."""
    base_dir = os.path.dirname(output_path)
    if base_dir:
        os.makedirs(base_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active

    if chat_name and chat_name.strip():
        sheet_name = chat_name.strip()[:31]
        for char in ['\\', '/', '*', '?', ':', '[', ']']:
            sheet_name = sheet_name.replace(char, '_')
        ws.title = sheet_name
    else:
        ws.title = f"Чат {chat_id}"[:31]

    styles = _create_styles()

    operations = db.get_operations(chat_id, limit=100000)
    _write_operations_tables_for_chat(ws, operations, styles, chat_id, db)

    wb.save(output_path)
    return output_path


def export_group_balances_to_excel(db: Database, filepath: str):
    """
    Экспорт таблицы остатков групп в Excel
    """
    table = db.get_group_balances_table()
    totals = db.get_total_balances_all_groups()
    currencies = list(CURRENCIES)

    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки групп"

    # ---------- Заголовок ----------
    headers = ["Группа"] + currencies
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font

    # ---------- Данные по группам ----------
    for group_name in sorted(table.keys()):
        safe_name = str(group_name).replace("\n", " ").replace("|", "/")
        row = [safe_name] + [table[group_name].get(cur, 0.0) for cur in currencies]
        ws.append(row)

    # ---------- ИТОГО ----------
    total_row_idx = ws.max_row + 1
    ws.append(["ИТОГО"] + [totals.get(cur, 0.0) for cur in currencies])

    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row_idx, column=col).font = Font(bold=True)

    # ---------- Форматирование ----------
    for col in range(2, len(headers) + 1):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "#,##0.00"

    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(filepath)

def _autosize(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 2, 45)

def export_report_income_matrix(rows, output_path: str, report_date: str):
    """
    rows: [(client_name, currency, amount, full_message), ...]
    full_message будет сохранён в комментарий к ячейке с клиентом.
    """
    agg = defaultdict(lambda: defaultdict(float))   # client -> currency -> sum
    totals = defaultdict(float)                     # currency -> sum
    comments = defaultdict(list)                    # client -> [full_message...]

    for client_name, cur, amt, full_msg in rows:
        agg[client_name][cur] += float(amt)
        totals[cur] += float(amt)
        if full_msg:
            comments[client_name].append(str(full_msg))

    currencies = sorted(totals.keys())
    clients = sorted(agg.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = f"Report_{report_date}"

    ws["A1"] = "Клиент"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    for j, cur in enumerate(currencies, start=2):
        c = ws.cell(row=1, column=j, value=cur)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    r = 2
    for client in clients:
        cell_a = ws.cell(row=r, column=1, value=client)

        # ✅ Комментарий к ячейке клиента
        if comments.get(client):
            note_text = "\\n\\n---\\n\\n".join(comments[client])

            # ограничение Excel на комментарии — подстрахуемся
            if len(note_text) > 30000:
                note_text = note_text[:30000] + "\\n\\n... (обрезано)"

            cell_a.comment = Comment(note_text, "report-bot")

        for j, cur in enumerate(currencies, start=2):
            v = agg[client].get(cur, 0.0)
            cell = ws.cell(row=r, column=j, value=v if v != 0 else "")
            cell.number_format = "#,##0.00"
        r += 1

    ws.cell(row=r, column=1, value="ИТОГО").font = Font(bold=True)
    for j, cur in enumerate(currencies, start=2):
        cell = ws.cell(row=r, column=j, value=totals.get(cur, 0.0))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"

    _autosize(ws)
    wb.save(output_path)


