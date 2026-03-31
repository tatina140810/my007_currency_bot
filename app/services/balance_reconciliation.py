import json
import logging
import datetime
from decimal import Decimal
import gspread
import openpyxl

from app.db.instance import db
from app.services.google_sheets import _get_gc, _execute_with_retry, _apply_time_patch
from app.core.config import CASSA_SPREADSHEET_ID

logger = logging.getLogger(__name__)

CURRENCIES_REPORT = ["Рубли", "USD", "Евро", "CNY", "тенге"]

def parse_balance_excel(file_path_or_wb, target_sheet_name: str = None) -> dict:
    """
    Парсит загруженный файл Остатки (Excel) или открытый workbook, и возвращает totals по валютам.
    Ожидаемые колонки: Рубли (B), USD (C), Евро (D), CNY (E), тенге/дирхам (F).
    """
    if isinstance(file_path_or_wb, str):
        wb = openpyxl.load_workbook(file_path_or_wb, data_only=True)
    else:
        wb = file_path_or_wb
        
    sheet = None
    if target_sheet_name:
        # Пробуем найти лист по точному или частичному совпадению без учета регистра
        for sname in wb.sheetnames:
            if target_sheet_name.lower() in sname.lower():
                sheet = wb[sname]
                logger.info(f"Нашли лист {sname} по запросу {target_sheet_name}")
                break
                
    if not sheet:
        sheet = wb.active
        logger.info(f"Используем активный лист {sheet.title}")
    
    totals = {
        "Рубли": 0.0,
        "USD": 0.0,
        "Евро": 0.0,
        "CNY": 0.0,
        "тенге": 0.0
    }
    
    # Ищем строку с ИТОГО (в колонке A или B)
    max_row = sheet.max_row
    
    # Оптимизация: не перебирать миллион строк, если лист пустой
    if max_row > 5000:
        logger.warning(f"Лист {sheet.title} имеет {max_row} строк, ограничиваем поиск.")
        max_row = min(max_row, 5000)
    
    found_row = None
    for r in range(max_row, 0, -1):
        cell_a = sheet.cell(row=r, column=1).value
        cell_b = sheet.cell(row=r, column=2).value
        text_a = str(cell_a).lower() if cell_a else ""
        text_b = str(cell_b).lower() if cell_b else ""
        
        if "итого" in text_a or "итого" in text_b:
            found_row = r
            break

    if not found_row:
        # Если не нашли слово ИТОГО, берем самую нижнюю непустую строку, где есть цифры
        for r in range(max_row, 0, -1):
            if isinstance(sheet.cell(row=r, column=3).value, (int, float)):
                found_row = r
                break

    if found_row:
        cols_mapping = {
            2: "Рубли",
            3: "USD",
            4: "Евро",
            5: "CNY",
            6: "тенге" # Может быть дирхам, но по задаче пишем "тенге" для колонки F
        }
        for col_idx, curr_name in cols_mapping.items():
            val = sheet.cell(row=found_row, column=col_idx).value
            try:
                totals[curr_name] = float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                totals[curr_name] = 0.0
                
    return totals

def _parse_date(date_str: str) -> datetime.date:
    try:
        return datetime.datetime.strptime(date_str, "%d.%m.%Y").date()
    except:
        return None

def fetch_daily_sums(target_date: datetime.date) -> dict:
    """
    Собирает суммы из листов:
    - ЗАПРОСЫ ПО ВХОД.СУММАМ И ДОКИ (приходы)
    - Платежи (расходы)
    - конветации (приходы/расходы)
    """
    _apply_time_patch()
    gc = _get_gc()
    sh = gc.open_by_key(CASSA_SPREADSHEET_ID)
    
    results = {
        "inputs": {"Рубли": 0.0, "USD": 0.0, "Евро": 0.0, "CNY": 0.0, "тенге": 0.0},
        "payments": {"Рубли": 0.0, "USD": 0.0, "Евро": 0.0, "CNY": 0.0, "тенге": 0.0},
        "swift_commissions": {"Рубли": 0.0, "USD": 0.0, "Евро": 0.0, "CNY": 0.0, "тенге": 0.0},
        "conv_income": {"Рубли": 0.0, "USD": 0.0, "Евро": 0.0, "CNY": 0.0, "тенге": 0.0},
        "conv_expense": {"Рубли": 0.0, "USD": 0.0, "Евро": 0.0, "CNY": 0.0, "тенге": 0.0},
        "zak_commissions": {"Рубли": 0.0, "USD": 0.0, "Евро": 0.0, "CNY": 0.0, "тенге": 0.0}
    }
    
    # 1. ЗАПРОСЫ ПО ВХОД.СУММАМ И ДОКИ
    try:
        ws_in = sh.worksheet("ЗАПРОСЫ ПО ВХОД.СУММАМ И ДОКИ")
        records = ws_in.get_all_values()
        for r in records[1:]:
            if len(r) >= 5:
                # Ожидается Дата (0 или 1), Валюта, Сумма. Пусть Дата в 0, Сумма в 3, Валюта в 4 (надо уточнить индексы)
                # Assuming generic scan for the date in the row
                row_str = " ".join(r)
                if target_date.strftime("%d.%m.%Y") in row_str:
                    # Попробуем найти сумму и валюту простым способом
                    # Обычно: Дата, Клиент, Банк, Сумма, Валюта
                    # Ищем подходящую валюту
                    val_curr = "Рубли"
                    val_sum = 0.0
                    for cell in r:
                        c_low = cell.lower()
                        if "usd" in c_low or "доллар" in c_low: val_curr = "USD"
                        elif "eur" in c_low or "евро" in c_low: val_curr = "Евро"
                        elif "cny" in c_low or "юань" in c_low: val_curr = "CNY"
                        elif "тенге" in c_low or "kzt" in c_low: val_curr = "тенге"
                        elif "руб" in c_low: val_curr = "Рубли"
                        
                        try:
                            # Пытаемся распарсить числа
                            clean_str = cell.replace(" ", "").replace(",", ".")
                            if clean_str.replace(".", "", 1).isdigit():
                                possible_sum = float(clean_str)
                                if possible_sum > val_sum: 
                                    val_sum = possible_sum
                        except:
                            pass
                    
                    results["inputs"][val_curr] += val_sum
    except Exception as e:
        logger.error(f"Error parsing ЗАПРОСЫ: {e}")

    # 2. Платежи
    try:
        ws_pay = sh.worksheet("Платежи")
        records = ws_pay.get_all_values()
        # Ожидаемый заголовки: "Отчет Back", "Компания", "Тип", "Контрагент", "Валюта платежа", "Сумма"
        in_target_date = False
        target_marker = f"--- ПЛАТЕЖИ ЗА {target_date.strftime('%d.%m.%Y')}"
        
        for r in records:
            row_str = " ".join(r)
            if target_marker in row_str:
                in_target_date = True
                continue
            if in_target_date and "--- ПЛАТЕЖИ ЗА" in row_str:
                in_target_date = False # Начался другой день
                
            if in_target_date and len(r) >= 6:
                curr_raw = r[4].lower()
                amount_raw = r[5].replace(" ", "").replace(",", ".")
                
                try:
                    swift_com = float(str(r[6]).replace(" ", "").replace(",", ".")) if len(r) >= 7 and str(r[6]).strip() else 0.0
                except:
                    swift_com = 0.0
                    
                try:
                    amt = float(amount_raw)
                    if "usd" in curr_raw or "доллар" in curr_raw: 
                        results["payments"]["USD"] += amt
                        results["swift_commissions"]["USD"] += swift_com
                    elif "eur" in curr_raw or "евро" in curr_raw: 
                        results["payments"]["Евро"] += amt
                        results["swift_commissions"]["Евро"] += swift_com
                    elif "cny" in curr_raw or "юань" in curr_raw: 
                        results["payments"]["CNY"] += amt
                        results["swift_commissions"]["CNY"] += swift_com
                    elif "тенге" in curr_raw or "kzt" in curr_raw: 
                        results["payments"]["тенге"] += amt
                        results["swift_commissions"]["тенге"] += swift_com
                    else: 
                        results["payments"]["Рубли"] += amt
                        results["swift_commissions"]["Рубли"] += swift_com
                except:
                    pass
    except Exception as e:
        logger.error(f"Error parsing Платежи: {e}")

    # 3. Конвертации
    try:
        ws_conv = sh.worksheet("конвертации")
        records = ws_conv.get_all_values()
        # Колонки: Дата(0), Клиент(1), Сумма(2), Валюта(3), Курс(4), Сумма РУБ(5)
        for r in records[1:]:
            if len(r) >= 6 and target_date.strftime("%d.%m.%Y") in r[0]:
                curr_raw = r[3].lower()
                amount_raw = r[2].replace(" ", "").replace(",", ".")
                rub_raw = r[5].replace(" ", "").replace(",", ".")
                try:
                    amt = float(amount_raw)
                    rub_amt = float(rub_raw)
                    
                    val_curr = None
                    if "usd" in curr_raw or "доллар" in curr_raw: val_curr = "USD"
                    elif "eur" in curr_raw or "евро" in curr_raw: val_curr = "Евро"
                    elif "cny" in curr_raw or "юань" in curr_raw: val_curr = "CNY"
                    elif "тенге" in curr_raw or "kzt" in curr_raw: val_curr = "тенге"
                    
                    if val_curr:
                        # Покупка инвалюты -> мы потратили рубли, получили инвалюту
                        results["conv_income"][val_curr] += amt
                        results["conv_expense"]["Рубли"] += rub_amt
                except:
                    pass
    except Exception as e:
        logger.error(f"Error parsing конветации: {e}")

    # 4. Процент снятия и пополнения (ZAK)
    try:
        ws_zak = sh.worksheet("Проценты_детально")
        records = ws_zak.get_all_values()
        for r in records[1:]:
            if len(r) >= 9 and target_date.strftime("%d.%m.%Y") in str(r[0]):
                curr_raw = str(r[4]).lower()
                fee_raw = str(r[8]).replace(" ", "").replace(",", ".")
                try:
                    fee = float(fee_raw)
                    if "usd" in curr_raw or "доллар" in curr_raw: results["zak_commissions"]["USD"] += fee
                    elif "eur" in curr_raw or "евро" in curr_raw: results["zak_commissions"]["Евро"] += fee
                    elif "cny" in curr_raw or "юань" in curr_raw: results["zak_commissions"]["CNY"] += fee
                    elif "тенге" in curr_raw or "kzt" in curr_raw: results["zak_commissions"]["тенге"] += fee
                    else: results["zak_commissions"]["Рубли"] += fee
                except:
                    pass
    except Exception as e:
        logger.error(f"Error parsing Проценты_детально: {e}")

    return results

def process_evening_reconciliation(date_str: str) -> str:
    """
    Вечерняя сверка: проверяет данные утра и вечера из DB и возвращает статус.
    Запись в Google Sheets выполняется отдельно через write_cash_report.py
    и _update_balance_in_sheet (уже вызваны в documents.py).
    """
    db_rec = db.get_daily_balance(date_str)
    if not db_rec:
        return f"Данные за {date_str} не найдены в базе."

    morning = json.loads(db_rec.get("morning_data") or "{}")
    evening = json.loads(db_rec.get("evening_data") or "{}")

    if not morning or not evening:
        return "Недостаточно данных (нет утра или вечера) для сверки."

    target_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    td_str = target_date.strftime("%d.%m.%Y")

    lines = [f"📊 Сверка за {td_str}:"]
    for curr in ["Рубли", "USD", "Евро", "CNY", "тенге"]:
        m = morning.get(curr, 0) or 0
        e = evening.get(curr, 0) or 0
        if m or e:
            lines.append(f"  {curr}: утро {m:,.2f} → вечер {e:,.2f}")

    lines.append("\n✅ Данные обновлены в листе «отчет по остаткам».")

    db.save_daily_balance(date_str, processed=True)
    return "\n".join(lines)

