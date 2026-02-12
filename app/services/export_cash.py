import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def export_cash_report(data: dict, output_path: str):
    """
    Generates Excel report for Cash Evening Report.
    data structure:
    {
        "date": "YYYY-MM-DD",
        "summary": {
            "USD": {"opening": 100, "deposit": 50, "withdraw": 20, "exchange_in": 0, "exchange_out": 10, "closing": 120},
            ...
        },
        "exchanges": [
            {"time": "HH:MM", "group": "Someone", "from_curr": "USD", "amount": 100, "to_curr": "RUB", "converted": 9000, "rate": 90, "user": "Manager"},
            ...
        ]
    }
    """
    wb = Workbook()
    
    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = f"Cash_Report_{data['date']}"
    
    headers = ["Currency", "Opening Balance", "Deposits", "Exchange +/-", "Withdrawals", "Closing Balance"]
    ws1.append(headers)
    
    # Styles
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Data Rows
    summary = data.get("summary", {})
    sorted_currencies = sorted(summary.keys())
    
    for cur in sorted_currencies:
        vals = summary[cur]
        # Exchange Net = In - Out
        exchange_net = vals.get("exchange_in", 0) - vals.get("exchange_out", 0)
        
        row = [
            cur,
            vals.get("opening", 0),
            vals.get("deposit", 0),
            exchange_net,
            vals.get("withdraw", 0),
            vals.get("closing", 0)
        ]
        ws1.append(row)

    # Farming formatting
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = "#,##0.00"

    # Autosize columns
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws1.column_dimensions[column].width = adjusted_width

    # --- Sheet 2: Exchanges ---
    ws2 = wb.create_sheet(f"Exchanges_{data['date']}")
    headers2 = ["Time", "Group", "From Currency", "Amount", "To Currency", "Converted Amount", "Internal Rate", "Comment/User"]
    ws2.append(headers2)
    
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    exchanges = data.get("exchanges", [])
    for ex in exchanges:
        row = [
            ex.get("time", ""),
            ex.get("group", ""),
            ex.get("from_curr", ""),
            ex.get("amount", 0),
            ex.get("to_curr", ""),
            ex.get("converted", 0),
            ex.get("rate", 0),
            ex.get("user", "")
        ]
        ws2.append(row)

    # Autosize columns sheet 2
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws2.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
